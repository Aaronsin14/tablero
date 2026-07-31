#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extractor del Tablero de EMBARQUE / PRODUCCION EN TIEMPO REAL - AngioDynamics.

Lee el archivo "Avance de embarque Diario.xlsx" (o el nombre que tenga la copia
en OneDrive) y genera data_embarque.json para el dashboard embarque_produccion.html.

Solo procesa las hojas cuyo nombre es una FECHA LIMPIA, ej:
    "04-Ago-2026", "11-Ago-2026", "23-Mar-2026", "14 abr-2026"
Se ignoran las hojas con sufijos (AEREO, Maritimo, "(Back order)", etc.) y
cualquier otra hoja (Planning, Backlog, Simulador Welder, Charts, etc.).

Estructura de cada hoja de fecha:
  Fila 2 = encabezados:  Exportation | Process | Sub ensamble | Nuevo Part Number
                         | Comentarios | Demanda Real | Delta | <procesos...>
  Fila 3+ = una fila por numero de parte / lote.
  Columnas de proceso (H en adelante): cada una es una ETAPA de la linea, en
  orden (Wip Welder -> Embedder -> Drawdown -> Sidehole -> Handrilling ->
  Baking -> Prepping -> Moldeo -> Hydrophilic Mariner -> Baking Mariner ->
  Pouch -> Boxing -> Packaged).

  Cada celda de proceso, cuando tiene numero, marca en que etapa esta
  actualmente ese numero de parte (la columna de proceso mas a la derecha que
  tenga dato = etapa actual). El TOTAL PRODUCIDO no se re-calcula: se toma
  directo de la columna "Total" (U), que el propio Excel ya calcula con
  formula. El "% Accomplisment" (V) del Excel tambien existe, pero aqui se
  recalcula como Total/Demanda para evitar los #DIV/0! que trae esa columna
  cuando la demanda esta vacia.

USO:
  python extract_embarque.py            (una vez)
  python extract_embarque.py --watch    (cada 15 min)
"""
import json
import re
import sys
import time
import datetime
from pathlib import Path

from openpyxl import load_workbook

# ----------------------------------------------------------------------------
# CONFIGURACION -- AJUSTA ESTA RUTA A TU ARCHIVO REAL
# Ruta confirmada con Get-Item (archivo real .xlsx, 4.29 MB, sincronizado).
# ----------------------------------------------------------------------------
EXCEL_PATH = Path(
    r"C:\Users\aaron.lara\OneDrive - Biomerics\BALA-CENTRAL - 1. Producción\%Avance de embarque Diario.xlsx"
)
PLANT_NAME = "AngioDynamics"
OUTPUT_JSON = Path(__file__).parent / "data_embarque.json"
REFRESH_SECONDS = 900  # 15 min

# Columnas fijas de identificacion (1-based)
COL_EXPORTATION = 1
COL_PROCESS = 2
COL_SUBENSAMBLE = 3
COL_PARTNUMBER = 4
COL_COMENTARIOS = 5
COL_DEMANDA = 6
COL_DELTA = 7
COL_PROC_START = 8   # primera columna de etapa de proceso (Wip Welder)
COL_PROC_END = 20    # ultima columna de etapa (Packaged)
COL_TOTAL = 21        # columna "Total" (U) -- ya calculada por el Excel
COL_PCT_ACC = 22      # columna "% Accomplisment" (V) -- ya calculada por el Excel
HEADER_ROW = 2
DATA_START_ROW = 3

MESES = {
    "ene": 1, "jan": 1, "feb": 2, "mar": 3, "abr": 4, "apr": 4,
    "may": 5, "jun": 6, "jul": 7, "ago": 8, "aug": 8,
    "sep": 9, "set": 9, "oct": 10, "nov": 11, "dic": 12, "dec": 12,
}
SHEET_DATE_RE = re.compile(
    r"^\s*(\d{1,2})[\s\-]+([A-Za-zÁÉÍÓÚáéíóúñÑ]{3,4})[\s\-]+(\d{4})\s*$"
)


def _clean(s):
    return re.sub(r"\s+", " ", str(s).strip())


def _num(v):
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if s == "" or s.startswith("#"):
            return None
        try:
            return float(s)
        except ValueError:
            return None
    if isinstance(v, (int, float)):
        return float(v)
    return None


def sheet_date(name):
    """Si el nombre de hoja es una fecha limpia 'DD-Mes-YYYY', regresa un date."""
    m = SHEET_DATE_RE.match(name)
    if not m:
        return None
    day, mon, year = m.groups()
    mon_key = mon.strip().lower()[:3]
    month = MESES.get(mon_key)
    if not month:
        return None
    try:
        return datetime.date(int(year), month, int(day))
    except ValueError:
        return None


def stage_key(label):
    """Genera una llave corta y estable a partir del encabezado del proceso."""
    base = re.sub(r"\d+\s*UPH", "", label, flags=re.I)
    base = re.sub(r"[^A-Za-z0-9]+", "_", base).strip("_").lower()
    return base or "etapa"


def read_stage_columns(ws):
    """Lee dinamicamente las columnas de proceso (H..T) desde la fila de encabezado."""
    stages = []
    for c in range(COL_PROC_START, COL_PROC_END + 1):
        raw = ws.cell(HEADER_ROW, c).value
        if not raw or not str(raw).strip():
            continue
        label = _clean(raw)
        uph_m = re.search(r"(\d+)\s*UPH", label, flags=re.I)
        stages.append({
            "col": c,
            "key": stage_key(label),
            "label": re.sub(r"\s*\d+\s*UPH", "", label, flags=re.I).strip(),
            "uph": int(uph_m.group(1)) if uph_m else None,
        })
    return stages


def read_sheet_parts(ws, stages):
    parts = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        pn = ws.cell(r, COL_PARTNUMBER).value
        if pn is None or str(pn).strip() == "":
            continue
        demanda = _num(ws.cell(r, COL_DEMANDA).value)

        # etapa actual = la columna de proceso (H..T) mas a la derecha con dato.
        # el TOTAL producido se toma directo de la columna "Total" (U) que ya
        # trae la formula del Excel, en vez de re-derivarlo nosotros.
        filled_stage = None
        for st in stages:
            v = _num(ws.cell(r, st["col"]).value)
            if v is not None and v > 0:
                filled_stage = st
        if filled_stage is None:
            continue  # numero de parte sin movimiento todavia -> no se muestra

        producido = _num(ws.cell(r, COL_TOTAL).value)
        if producido is None:
            producido = _num(ws.cell(r, filled_stage["col"]).value)  # respaldo

        pct = (producido / demanda * 100.0) if demanda else None
        if pct is None:
            status = "sin_meta"
        elif pct >= 100:
            status = "verde"
        elif pct >= 80:
            status = "ambar"
        else:
            status = "rojo"

        parts.append({
            "part_number": _clean(pn),
            "sub_ensamble": _clean(ws.cell(r, COL_SUBENSAMBLE).value) if ws.cell(r, COL_SUBENSAMBLE).value else None,
            "demanda": demanda,
            "producido": producido,
            "pct": round(pct, 1) if pct is not None else None,
            "stage_key": filled_stage["key"],
            "stage_col": filled_stage["col"],
            "status": status,
        })
    return parts


def find_target_sheets(wb):
    """Regresa lista de (sheet_name, date) para hojas con nombre = fecha limpia."""
    out = []
    for name in wb.sheetnames:
        d = sheet_date(name)
        if d:
            out.append((name, d))
    return out


def build_payload():
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"No se encontro el Excel en: {EXCEL_PATH}")

    wb = load_workbook(EXCEL_PATH, data_only=True)
    targets = find_target_sheets(wb)
    if not targets:
        raise KeyError("No se encontraron hojas con nombre de fecha limpia (ej. '04-Ago-2026').")

    hoy = datetime.date.today()
    # la semana "actual" = el PROXIMO embarque (fecha mas cercana >= hoy, es
    # decir: el dia mismo de la fecha AUN muestra esa hoja; hasta el dia
    # SIGUIENTE brinca a la siguiente fecha). Si no hay ninguna fecha futura
    # o de hoy, se usa la mas reciente ya pasada.
    futuras = sorted([t for t in targets if t[1] >= hoy], key=lambda t: t[1])
    if futuras:
        current_name, current_date = futuras[0]
    else:
        pasadas = sorted(targets, key=lambda t: t[1])
        current_name, current_date = pasadas[-1]
    ws = wb[current_name]

    stages = read_stage_columns(ws)
    parts = read_sheet_parts(ws, stages)

    total_demanda = sum(p["demanda"] for p in parts if p["demanda"])
    total_producido = sum(p["producido"] for p in parts if p["producido"])
    # % Cumplimiento General = solo lo que YA LLEGO a Packaged cuenta como
    # cumplido (termino todo su recorrido). Lo que este mas avanzado pero aun
    # no llegue a Packaged (Baking, Moldeo, Pouch, etc.) NO suma aqui todavia
    # -- sigue siendo trabajo en proceso, no cumplimiento.
    empacado = sum(p["producido"] for p in parts if p["stage_key"] == "packaged" and p["producido"])
    pct_general = (empacado / total_demanda * 100.0) if total_demanda else 0.0
    atrasados = sum(1 for p in parts if p["status"] == "rojo")

    weeks_available = sorted(
        [{"label": n, "date": d.isoformat()} for n, d in targets],
        key=lambda w: w["date"],
    )

    payload = {
        "plant": PLANT_NAME,
        "source_file": EXCEL_PATH.name,
        "generated_at": datetime.datetime.now().isoformat(timespec="seconds"),
        "current_sheet": current_name,
        "current_date": current_date.isoformat(),
        "stages": [{"key": s["key"], "label": s["label"], "uph": s["uph"]} for s in stages],
        "parts": parts,
        "kpis": {
            "partes_activos": len(parts),
            "total_demanda": total_demanda,
            "total_producido": total_producido,
            "empacado": empacado,
            "pct_cumplimiento": round(pct_general, 1),
            "exceso_total": round(total_producido - total_demanda, 0),
            "atrasados": atrasados,
        },
        "weeks_available": weeks_available,
    }
    return payload


def write_json():
    payload = build_payload()
    OUTPUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"[{datetime.datetime.now():%H:%M:%S}] data_embarque.json: "
          f"hoja '{payload['current_sheet']}', {payload['kpis']['partes_activos']} partes activas")


def main():
    watch = "--watch" in sys.argv
    while True:
        try:
            write_json()
        except Exception as e:
            print(f"[ERROR] {e}", file=sys.stderr)
        if not watch:
            break
        time.sleep(REFRESH_SECONDS)


if __name__ == "__main__":
    main()
