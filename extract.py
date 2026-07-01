#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extractor del Tablero Tier 2 - AngioDynamics.

Lee la SECCION 'Angiodynamics' dentro de la hoja 'Tier 2' del archivo
Legacy Production Performance 2026.xlsx y genera data.json con TODAS las
semanas (no se salta ninguna).

USO MANUAL:      python extract.py
USO AUTOMATICO:  python extract.py --watch      (cada 1 hora)
"""

import json
import sys
import time
import datetime
from pathlib import Path

from openpyxl import load_workbook

# ----------------------------------------------------------------------------
# CONFIGURACION  --  AJUSTA ESTA RUTA A TU ARCHIVO REAL
# ----------------------------------------------------------------------------
EXCEL_PATH = Path(
    r"C:\Users\aaron.lara\OneDrive - Biomerics\BEX BALA - Legacy Production Performance 2026\Legacy Production Performance 2026.xlsx"
)
SHEET_NAME = "Tier 2"
PLANT_NAME = "AngioDynamics"
OUTPUT_JSON = Path(__file__).parent / "data.json"
REFRESH_SECONDS = 3600  # 1 hora

# Para pruebas locales con el archivo de muestra:
# EXCEL_PATH = Path("Legacy_Production_Performance_2026.xlsx")

SECTION_HEADER_TEXT = "angiodynamics"   # se busca en columna A (case-insensitive)

KPI_DEFS = [
    ("safety",       "Safety (Events)",          "count"),
    ("quality",      "Quality (Events) MRB",     "count"),
    ("daily_ins",    "Daily Ins (units)",        "units"),
    ("daily_outs",   "Daily Outs (FG)",          "units"),
    ("yield",        "Yield (%)",                "pct"),
    ("scrap",        "SCRAP",                    "pct"),
    ("downtime",     "Unplanned downtime (hrs)", "hours"),
    ("attrition",    "Attrition (Bottle neck)",  "count"),
    ("productivity", "Productivity",             "pct"),
]


def _num(v):
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if s == "" or s.startswith("#"):
            return None
        try:
            return float(s)
        except ValueError:
            return None
    if isinstance(v, (int, float)):
        return float(v)
    return None


def find_section(ws):
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip().lower() == SECTION_HEADER_TEXT:
            # validar que sea la SECCION real: r+1 debe decir "ScoreCard"
            nxt = ws.cell(r + 1, 1).value
            if not (isinstance(nxt, str) and "scorecard" in nxt.strip().lower()):
                continue
            header_row = r + 1      # "Week NN" labels
            date_row   = r + 2      # fechas + "Week NN" en columna total
            day_row    = r + 3      # "KPI / Day 1..7"
            kpi_base   = r + 4      # primera fila de KPI (Safety Planned)
            return header_row, date_row, day_row, kpi_base
    raise KeyError(f"No se encontro la seccion '{PLANT_NAME}' en la hoja '{SHEET_NAME}'.")


def detect_week_blocks(ws, header_row, date_row):
    blocks = []
    for c in range(1, ws.max_column + 1):
        lbl = ws.cell(header_row, c).value
        if isinstance(lbl, str) and lbl.strip().lower().startswith("week"):
            day_cols = list(range(c, c + 7))
            total_col = None
            for tc in range(c + 7, min(c + 10, ws.max_column + 1)):
                dv = ws.cell(date_row, tc).value
                if isinstance(dv, str) and dv.strip().lower().startswith("week"):
                    total_col = tc
                    break
            if total_col is None:
                total_col = c + 7
            blocks.append({
                "label": lbl.strip(),
                "start_col": c,
                "day_cols": day_cols,
                "total_col": total_col,
            })
    return blocks


def build_payload():
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"No se encontro el Excel en: {EXCEL_PATH}")

    wb = load_workbook(EXCEL_PATH, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise KeyError(f"La hoja '{SHEET_NAME}' no existe. Hojas: {wb.sheetnames}")
    ws = wb[SHEET_NAME]

    header_row, date_row, day_row, kpi_base = find_section(ws)
    blocks = detect_week_blocks(ws, header_row, date_row)

    kpi_rows = {key: kpi_base + i * 3 for i, (key, _, _) in enumerate(KPI_DEFS)}

    weeks = {}
    week_order = []
    seen = {}
    for blk in blocks:
        base_key = blk["label"].replace(" ", "_")
        seen[base_key] = seen.get(base_key, 0) + 1
        wkey = base_key if seen[base_key] == 1 else f"{base_key}_{seen[base_key]}"
        week_order.append(wkey)

        day_cols = blk["day_cols"]
        total_col = blk["total_col"]

        dates = []
        for c in day_cols:
            v = ws.cell(date_row, c).value
            dates.append(v.strftime("%Y-%m-%d") if isinstance(v, datetime.datetime) else None)

        kpis = {}
        for key, label, ktype in KPI_DEFS:
            base = kpi_rows[key]
            kpis[key] = {
                "label": label,
                "type": ktype,
                "planned": [_num(ws.cell(base,     c).value) for c in day_cols],
                "actual":  [_num(ws.cell(base + 1, c).value) for c in day_cols],
                "diff":    [_num(ws.cell(base + 2, c).value) for c in day_cols],
                "planned_total": _num(ws.cell(base,     total_col).value),
                "actual_total":  _num(ws.cell(base + 1, total_col).value),
                "diff_total":    _num(ws.cell(base + 2, total_col).value),
            }

        weeks[wkey] = {"label": blk["label"], "dates": dates, "kpis": kpis}

    payload = {
        "plant": PLANT_NAME,
        "source_file": EXCEL_PATH.name,
        "generated_at": datetime.datetime.now().isoformat(timespec="seconds"),
        "weeks": weeks,
        "week_order": week_order,
        "week_count": len(week_order),
    }
    return payload


def write_json():
    payload = build_payload()
    OUTPUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[{datetime.datetime.now():%H:%M:%S}] data.json actualizado: "
          f"{payload['week_count']} semanas desde {payload['source_file']}")


def main():
    watch = "--watch" in sys.argv
    while True:
        try:
            write_json()
        except Exception as e:
            print(f"[ERROR] {e}", file=sys.stderr)
        if not watch:
            break
        time.sleep(REFRESH_SECONDS)


if __name__ == "__main__":
    main()
