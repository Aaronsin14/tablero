#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extractor del dashboard HORA A HORA - AngioDynamics (v4).
Lee la tabla de MAQUINAS (valores reales, no formulas) y agrupa por proceso.

Por que la tabla de maquinas y no la resumida:
  La tabla resumida usa formulas (=G307+G308). Python solo ve el valor cacheado,
  que puede venir vacio si Excel no recalculo antes de guardar. La tabla de
  maquinas tiene numeros escritos directo, asi que siempre es confiable.

Estructura (por dia, encabezados en fila dia+4):
  Turno A: proc B(2),  metaDia C(3),  UPH D(4),  horas G(7)..    10 horas
  Turno B: proc AH(34),metaDia AI(35),UPH AJ(36),horas AM(39)..   7 horas
  Turno C: proc BH(60),metaDia BI(61),UPH BJ(62),horas BM(65)..   8 horas
  Cada hora ocupa 2 columnas: Actual + Scrap
  DELTA y DOWN TIME al final de cada turno.

Si el UPH cambia en el Excel, la Meta*Dia tambien cambia; ambos se leen directo,
asi que el dashboard siempre refleja los valores actuales.
"""
import json, sys, time, datetime, re
from pathlib import Path
from openpyxl import load_workbook

EXCEL_PATH = Path(
    r"C:\Users\aaron.lara\OneDrive - Biomerics\BALA-CENTRAL - Hr a Hr actulizado 2026\Control_hr-hr_angio_2026_Actulizado.xlsx"
)
OUTPUT_JSON = Path(__file__).parent / "data_hora.json"
REFRESH_SECONDS = 900  # 15 minutos

DIAS = ["LUNES","MARTES","MIERCOLES","JUEVES","VIERNES","SABADO","DOMINGO"]
HDR_OFFSET = 4

# turno: nombre, col_proc, col_metaDia, col_uph, col_hora0, n_horas, col_delta, col_downtime
TURNOS = [
    ("A", 2,  3,  4,  7,  10, 30, 31),
    ("B", 34, 35, 36, 39, 7,  56, 57),
    ("C", 60, 61, 62, 65, 8,  84, 85),
]

def _num(v):
    return float(v) if isinstance(v,(int,float)) else None

def clean(s): return re.sub(r'\s+',' ',str(s).strip())

def group_name(proc):
    """'Cut (Tip/Teflon) 1' -> 'Cut'; 'Welder 3' -> 'Welder'."""
    p = re.sub(r'\s*\d+\s*$', '', str(proc).strip())
    p = re.sub(r'\(.*?\)', '', p).strip()
    p = re.sub(r'\s+(Soft|Accu)\s*Vu.*$', '', p).strip()
    return p if p else clean(proc)

def find_day_rows(ws):
    rows={}
    for r in range(1, ws.max_row+1):
        v=ws.cell(r,1).value
        if isinstance(v,str):
            up=v.strip().upper()
            for d in DIAS:
                if up.startswith(d) and d not in rows: rows[d]=r
    return rows

def find_summary_row(ws, day_row, next_day_row):
    """Fila donde empieza la tabla RESUMIDA de este dia (tiene 'Cantidad de estaciones').
    Leemos solo la tabla de maquinas, que esta ANTES de esta fila."""
    fin = next_day_row if next_day_row else ws.max_row
    for r in range(day_row, fin):
        for c in (4, 36, 62):   # col D / AJ / BJ segun turno
            v=ws.cell(r,c).value
            if isinstance(v,str) and 'cantidad' in v.lower():
                return r
    return None

def read_turno(ws, day_row, tcfg, next_day_row=None, summary_row=None):
    name,cproc,cmetaDia,cuph,chour0,nhours,cdelta,cdown = tcfg
    hdr = day_row + HDR_OFFSET
    hour_labels=[]
    for hi in range(nhours):
        c=chour0+hi*2
        hl=ws.cell(hdr-1,c).value
        hour_labels.append(clean(hl) if hl else f"H{hi+1}")

    groups={}; order=[]
    # el bloque de un dia puede tener varias sub-tablas (Fase 1, Fase 2...) separadas
    # por filas vacias y encabezados 'Proceso' repetidos. Recorremos hasta el
    # siguiente dia y saltamos vacios/encabezados en vez de cortar.
    # leer solo hasta ANTES de la tabla resumida (que duplica los datos)
    if summary_row:
        limite = summary_row - 1
    elif next_day_row:
        limite = next_day_row - 1
    else:
        limite = hdr + 95
    vacias = 0
    for r in range(hdr+1, limite):
        proc=ws.cell(r,cproc).value
        # fila sin proceso: puede ser separador entre fases -> saltar
        if not proc or not str(proc).strip():
            vacias += 1
            if vacias > 12: break   # muchas vacias seguidas = fin del dia
            continue
        vacias = 0
        # encabezado repetido de una nueva fase -> saltar
        if str(proc).strip() in ("Proceso","PROCESO"): continue
        # si aparece el nombre de otro dia, terminamos
        if any(str(proc).strip().upper().startswith(d) for d in DIAS): break

        metaDia=_num(ws.cell(r,cmetaDia).value)
        uph=_num(ws.cell(r,cuph).value)
        if metaDia is None and uph is None: continue

        g=group_name(proc)
        if g not in groups:
            groups[g]={"name":g,"maquinas":0,"meta_dia":0.0,"meta_hora":0.0,
                       "hours":[0.0]*nhours,"scrap":0.0,"delta":0.0,"downtime":0.0}
            order.append(g)
        G=groups[g]
        G["maquinas"]+=1
        G["meta_dia"]+=metaDia or 0
        G["meta_hora"]+=uph or 0
        for hi in range(nhours):
            c=chour0+hi*2
            G["hours"][hi]+=_num(ws.cell(r,c).value) or 0
            G["scrap"]  += _num(ws.cell(r,c+1).value) or 0
        G["downtime"]+= _num(ws.cell(r,cdown).value) or 0

    # DELTA = meta del dia - lo producido (positivo = falta; negativo = se paso)
    procesos=[]
    for g in order:
        G=groups[g]
        producido=sum(G["hours"])
        G["delta"]=round(G["meta_dia"]-producido, 2)
        procesos.append(G)
    return {"hour_labels":hour_labels,"procesos":procesos}

def build_payload():
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"No se encontro el Excel en: {EXCEL_PATH}")
    wb=load_workbook(EXCEL_PATH, data_only=True)
    weeks={}; week_order=[]
    for sh in wb.sheetnames:
        key=sh.strip()
        if not key.upper().startswith("WW"): continue
        ws=wb[sh]
        day_rows=find_day_rows(ws)
        # ordenar los dias por su fila para saber donde termina cada uno
        ordenados = sorted(day_rows.items(), key=lambda kv: kv[1])
        siguiente = {}
        for i,(d,fila) in enumerate(ordenados):
            siguiente[d] = ordenados[i+1][1] if i+1 < len(ordenados) else None
        dias={}
        for d in DIAS:
            if d not in day_rows: continue
            sr = find_summary_row(ws, day_rows[d], siguiente[d])
            turnos={}
            for tcfg in TURNOS:
                turnos[tcfg[0]]=read_turno(ws, day_rows[d], tcfg, siguiente[d], sr)
            dias[d]=turnos
        weeks[key]={"label":key,"dias":dias}
        week_order.append(key)
    return {
        "plant":"AngioDynamics",
        "source_file":EXCEL_PATH.name,
        "generated_at":datetime.datetime.now().isoformat(timespec="seconds"),
        "weeks":weeks,"week_order":week_order,"dias_orden":DIAS,
    }

def write_json():
    p=build_payload()
    OUTPUT_JSON.write_text(json.dumps(p,ensure_ascii=False,indent=1),encoding="utf-8")
    print(f"[{datetime.datetime.now():%H:%M:%S}] data_hora.json: {len(p['week_order'])} semanas")

def main():
    watch="--watch" in sys.argv
    while True:
        try: write_json()
        except Exception as e: print(f"[ERROR] {e}",file=sys.stderr)
        if not watch: break
        time.sleep(REFRESH_SECONDS)

if __name__=="__main__": main()
